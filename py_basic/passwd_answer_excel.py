import wx
from socket import socket,AF_INET,SOCK_STREAM
import time
from threading import Thread
import openpyxl

#加载xlsx文件----------------------------------------------------------------------------------------------------------------------------
work_book = openpyxl.load_workbook(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")
work_sheet = work_book.active
#创建服务器类------------------------------------------------------------------------------------------------------------------------------
class server_answer(wx.Frame):
    def __init__(self,work_sheet,work_book):
        self.isON = False
        self.host_port = ('',8888)
        self.work_sheet = work_sheet
        self.socket_connecting = {}
        self.work_book = work_book

 #绘制服务器窗口（finish）------------------------------------------------------------------------------------------------------------------
        wx.Frame.__init__(self,None,id=1222,pos=wx.DefaultPosition,title='英雄库',size=(550,550))
        pl = wx.Panel(self)
        box = wx.BoxSizer(wx.VERTICAL)
        ff = wx.FlexGridSizer(wx.HSCROLL)
        server_btn = wx.Button(pl,size=(200,40),label='开始服务')
        stop_server_btn = wx.Button(pl,size=(200,40),label='停止服务')
        ff.Add(server_btn,1,wx.TOP|wx.ALL,10)
        ff.Add(stop_server_btn,1,wx.TOP|wx.ALL,10)
        self.read_text = wx.TextCtrl(pl,size=(500,450),style=wx.TE_MULTILINE|wx.TE_READONLY)
        box.Add(ff,1,wx.ALIGN_CENTRE|wx.ALL,10)
        box.Add(self.read_text,1,wx.ALIGN_CENTRE|wx.ALL,10)
        pl.SetSizer(box)
 #绑定事件(finish)-------------------------------------------------------------------------------------------------------------------------
        self.Bind(wx.EVT_BUTTON,self.start_server,server_btn)
        self.Bind(wx.EVT_BUTTON,self.stop_server,stop_server_btn)
 #定义事件()-------------------------------------------------------------------------------------------------------------------------
    def send_message(self,socket,data):
        socket.send(data.encode('utf-8'))
    def recv_message(self,socket):
        return socket.recv(1024).decode('utf-8')
 #开始服务事件(finish)
    def start_server(self,event):
        if self.isON == False:
            self.isON = True
            self.read_text.AppendText("服务开启\n" + time.strftime('%Y-%m-%d %H:%M-%S', time.localtime()) + '\n')
            main_thread = Thread(target=self.start_work)
            main_thread.daemon = True
            main_thread.start()
 #主进程事件()
    def start_work(self):
        self.listen_socket = socket(AF_INET, SOCK_STREAM)
        self.listen_socket.bind(self.host_port)
        self.listen_socket.listen(5)
        while self.isON:
            try:
                session_socket,client_addr = self.listen_socket.accept()#接收一个套接字
                self.socket_connecting[session_socket] = ''
                session_Thread = session_thread(session_socket,self,self.work_sheet,self.work_book)
                session_Thread.daemon = True
                session_Thread.start()
            except OSError:
                break


    def stop_server(self,event):
        for session_socket in self.socket_connecting:
            session_socket.send("exit".encode('utf-8'))
            data = session_socket.recv(1024).decode('utf-8')
            if not data:
                session_socket.close()
            del self.socket_connecting[session_socket]
        if not self.socket_connecting:
            self.isON = False
            self.listen_socket.close()

    def root_stop(self):
        for session_socket in self.socket_connecting:
            session_socket.send("exit".encode('utf-8'))
            data = session_socket.recv(1024).decode('utf-8')
            if not data:
                session_socket.close()
            del self.socket_connecting[session_socket]
        if not self.socket_connecting:
            self.isON = False
            self.listen_socket.close()

    def stop_user(self,user):
        for session_socket in self.socket_connecting:
            if self.socket_connecting[session_socket] == user:
                session_socket.send("exit".encode('utf-8'))
                data = session_socket.recv(1024).decode('utf-8')
                if not data:
                    session_socket.close()
                    del self.socket_connecting[session_socket]

    def change_condition(self,user):



class session_thread(Thread):
    def __init__(self,session_socket,server,work_sheet,work_book):
        super().__init__()
        self.session_socket = session_socket
        self.server = server
        self.work_sheet = work_sheet
        self.work_book = work_book
        self.isON = True

    def run(self) -> None:
        while self.isON:
            self.send_message("欢迎使用本系统！")
            method = self.recv_message()
            if method == 'login':
                while True:
                    user_qx,user_name= self.user_varify()
                    if user_name != 'N':
                        break
                #root用户
                if user_qx == 'admin':
                    self.root_ctrl(user_name)
                #高权限用户
                elif user_qx == 'common':
                    self.common_ctrl(user_name)
                #普通用户
                elif user_qx == 'sudo':
                    self.sudo_ctrl(user_name)
            elif method == 'sign':
                self.sign_user()
            else:
                self.send_message("请先登录系统！")
    #不同用户操作----------------------------------------------------------------------------------------------------------
    #root操作
    def root_ctrl(self,name):
        while True:
            option = "查询信息\n更改权限\n强制下线\n关闭系统"
            self.send_message(option)
            option_list = option.split("\n")
            order = self.recv_message()
            while order not in option_list:
                self.send_message("无效指令")
                order = self.recv_message()
            if order == option_list[0]:
                option = "查询在线用户\n查询用户信息"
                option_list = option.split("\n")
                order = self.recv_message()
                while order not in option_list:
                    self.send_message("无效指令")
                    order = self.recv_message()
                if order == option_list[0]:
                    self.ro0_0()
                elif order == option_list[1]:
                    self.ro0_1()
            elif order == option_list[1]:
                self.ro1()
            elif order == option_list[2]:
                self.ro2()
            elif order == option_list[3]:
                server.root_stop()
    #查询所有用户的所有信息
    def ro0_0(self):
        rows = self.work_sheet.max_row +1
        user_online = ''
        for i in range(2,rows):
            if self.work_sheet.cell(i,8).value == 'on':
                user_online += f"{self.work_sheet.cell(i,1).value}\n"
        self.send_message(user_online)
    def ro0_1(self):
        rows = rows = self.work_sheet.max_row +1
        user_list = []
        send_data = ''
        for i in range(2,rows):
            user_list.append(self.work_sheet.cell(i,1).value)
            send_data += f"{self.work_sheet.cell(i,1).value}\n"
        self.send_message(send_data)
        order = self.recv_message()
        while order not in send_data:
            self.send_message("无效指令")
            order = self.recv_message()
        user_row = user_list.index(order)+2
        send_data = ''
        for i in range(1,11):
            send_data += f"{self.work_sheet.cell(1,i).value}   :   {self.work_sheet.cell(user_row,i).value}\n"
        self.send_message(send_data)

    def ro1(self):
        rows = self.work_sheet.max_row + 1
        user_list = []
        send_data = ''
        for i in range(2,rows):
            send_data += f"{self.work_sheet.cell(i,1).value    :     {self.work_sheet.cell(i,3)}\n}"
            user_list.append(self.work_sheet.cell(i,1).value)
        self.send_message(send_data)
        order = self.recv_message()
        while order not in user_list:
            self.send_message("无效指令")
            order = self.recv_message()
        user_row = user_list.index(order) + 2
        self.work_sheet.cell(user_row,3).value = self.change_qx(self.work_sheet.cell(user_row,3).value)
        self.send_message("权限已更改！")

    def change_qx(self,qx):
        if qx == "sudo":
            return common
        else:
            return sudo

    def ro2(self):
        rows = self.work_sheet.max_row +1
        user_online = ''
        for i in range(2,rows):
            if self.work_sheet.cell(i,8).value == 'on':
                user_online += f"{self.work_sheet.cell(i,1).value}\n"
        self.send_message(user_online)
        order = self.recv_message()
        while order not in user_online:
            self.send_message("无效指令")
            order = self.recv_message()
        server.stop_user(order)
        self.send_message("已让该用户下线！")


    #sudo用户操作
    def sudo_ctrl(self,name):
        rows = self.work_sheet.max_row + 1
        for i in range(2, rows):
            if self.work_sheet.cell(i, 1).value == name:
                user_row = i
                break
        while True:
            option = "查询信息\n打印本月财务报表"
            option_list = option.split("\n")
            order = self.recv_message()
            while order not in option_list:
                self.send_message("无效指令\n请重新输入")
                order = self.recv_message()
            if order == option_list[0]:
                option = "查询本用户信息\n查询common用户密码"
                option_list = option.split("\n")
                while order not in option_list:
                    self.send_message("无效指令\n请重新输入")
                    order = self.recv_message()
                if order == option_list[0]:
                    self.so0_1(user_row)
                if order == option_list[1]:
                    self.so0_2(rows)
            elif order == option_list[1]:
                self.so1(user_row)

    def so0_1(self,row):
        option_dict = {}
        for i in range(1,11):
            option_dict[self.work_sheet.cell(1,i).value] = self.work_sheet.cell(row,i).value
        order = self.recv_message()
        while order not in option_dict.keys():
            self.send_message("无效指令\n请重新输入")
            self.recv_message()
        self.send_message(option_dict[order])
    def so0_2(self,rows):
        com_passwd_dict = {}
        send_data = ''
        for i in range(3,rows):
            if self.work_sheet.cell(i,3).value == 'common':
                com_passwd_dict[self.work_sheet.cell(i,1).value] = self.work_sheet.cell(i,2).value
                send_data += f"{self.work_sheet.cell(i,1).value}\n"
        self.send_message(send_data)
        order = self.recv_message()
        while order not in send_data:
            self.send_message("无效指令\n请重新输入")
            self.recv_message()
        self.send_message(f"{order}  的密码：  {com_passwd_dict[order]}")
    def so1(self,row):
        self.send_message(self.work_sheet.cell(row,9).value)


    #普通用户操作
    def common_ctrl(self,name):
        rows  = self.work_sheet.max_row + 1
        for i in range(2,rows):
            if self.work_sheet.cell(i,1).value == name:
                user_row = i
                break
        while True:
            option = "查询信息\n打印本月财务报表"
            option_list =option.split("\n")
            self.send_message(option)
            order = self.recv_message()
            if order not in option_list:
                self.send_message("无效指令")
            elif order == option_list[0]:
                self.co0(user_row)
            elif order == option_list[1]:
                self.co1(user_row)


    def co0(self,row):
        option_dict = {}
        for i in range(1,8):
            option_dict[self.work_sheet.cell(1,i).value] = self.work_sheet.cell(row,i)
        order = self.recv_message()
        while order not in option_dict.keys():
            self.send_message("无效指令\n请重新输入")
            self.recv_message()
        self.send_message(option_dict[order])
    def co1(self,row):
        self.send_message(self.work_sheet.cell(row,9).value)


    def sign_user(self):
        rows = self.work_sheet.max_row + 1
        user_dict = []
        for i in range(3,rows):
            user_dict.append(self.work_sheet.cell(i,1).value)
        self.send_message("请输入注册的用户名")
        user_name = self.recv_message()
        while user_name in user_dict:
            self.send_message("该用户已存在！")
            user_name = self.recv_message()
        self.send_message("请输入密码")
        passwd = self.recv_message()
        while passwd == '':
            self.send_message("密码不能为空！")
            passwd = self.recv_message()
        self.send_message("请输入性别")
        sex = self.recv_message()
        while sex != '男' or sex != '女' or sex != 'female' or sex != 'male':
            self.send_message("只能填写（男/女/female/male）")
            sex = self.recv_message()
        self.send_message("请输入居住地")
        address = self.recv_message()
        while address == '':
            self.send_message("居住地不能为空！")
            address = self.recv_message()
        self.send_message("请输入生日（1111-11-11）")
        date = self.recv_message()
        while date == '':
            self.send_message("生日不能为空！")
            date = self.recv_message()
        qx = 'common'
        id = self.work_sheet.max_row
        bb = 'dads'
        user_cond = 'down'
        sign_time = time.strftime("%Y-%m-%d",time.localtime())
        new_user_list =[user_name,passwd,qx,sex,date,address,sign_time,user_cond,bb,id]
        for i in range(1,11):
            self.work_sheet.cell(rows,i).value = new_user_list[i-1]
        self.send_message("用户已创建！")





    #定义简单函数----------------------------------------------------------------------------------------------------------
    # 检测用户名，如果正确就返回用户名与权限
    def user_login_varify(self):
        user_name = self.recv_message()
        passwd = int(self.recv_message())
        user_passwd_dict = {}
        user_qx_dict = {}
        user_rows = self.work_sheet.max_row
        for i in range(2, user_rows + 1):
            user_passwd_dict[self.work_sheet.cell(i, 1).value] = self.work_sheet.cell(i, 2).value  # 数字类型都是int
            user_qx_dict[self.work_sheet.cell(i, 1).value] = self.work_sheet.cell(i, 3).value
        if user_name in user_passwd_dict.keys():
            if passwd == user_passwd_dict[user_name]:
                self.send_message("登录成功！")
                server.read_text.AppendText(
                    f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{user_name}进入了系统\n")
                return user_qx_dict[user_name], user_name
            else:
                self.send_message("密码错误！")
                return N, N
        else:
            self.send_message("用户不存在！")
            return N, N
    #定义接受信息的函数
    def recv_message(self):
        return self.session_socket.recv(1024).decode('utf-8')
    #定义发信息的函数
    def send_message(self,data):
        self.session_socket.send(data.encode('utf-8'))

#主程序运行---------------------------------------------------------------------------------------------------------------
if __name__ == '__main__':
    app = wx.App()
    server = server_answer(work_sheet,work_book)
    server.Show()
    app.MainLoop()
