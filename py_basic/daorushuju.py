import openpyxl

# wb = openpyxl.Workbook()
# ws = wb.active
# for i in range(1,161):
#     ws[f'A{i}']=0.5*i
# wb.save("物理数据.xlsx")

wb = openpyxl.load_workbook(r'D:\PyCharm Community Edition 2023.3\py_study\py_basic\物理数据.xlsx')
ws= wb.active
# i= ws.max_row
# print(i)
print("输入值")
for i in range(1,ws.max_row+1):
    value = input(f"{0.5*i}:")
    while value is None:
        ws[f'B{i}'] = value
        value = input(f"{0.5 * i}:")
wb.save(r'D:\PyCharm Community Edition 2023.3\py_study\py_basic\物理数据.xlsx')