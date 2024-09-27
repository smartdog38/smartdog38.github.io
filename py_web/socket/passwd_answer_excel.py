import wx
from socket import socket,AF_INET,SOCK_STREAM
import time
from threading import Thread
import openpyxl

#加载xlsx文件-------------------------------------------------------------------------------------------------------------
work_book = openpyxl.load_workbook(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")
work_sheet = work_book.active
#创建服务器类-------------------------------------------------------------------------------------------------------------
class server_answer(wx.Frame):
    def __init__(self,work_sheet,work_book):
        self.isON = False
        self.host_port = ('',8888)
        self.work_sheet = work_sheet
        self.socket_connecting = {}
        self.work_book = work_book
 #绘制服务器窗口（finish）--------------------------------------------------------------------------------------------------
        wx.Frame.__init__(self,None,id=1222,pos=wx.DefaultPosition,title='英雄库',size=(550,550))
        pl = wx.Panel(self)
        box = wx.BoxSizer(wx.VERTICAL)
        ff = wx.FlexGridSizer(wx.HSCROLL)
        server_btn = wx.Button(pl,size=(200,40),label='开始服务')
        stop_server_btn = wx.Button(pl,size=(200,40),label='停止服务')
        ff.Add(server_btn,1,wx.TOP|wx.ALL,10)
        ff.Add(stop_server_btn,1,wx.TOP|wx.ALL,10)
        self.read_text = wx.TextCtrl(pl,size=(500,450),style=wx.TE_MULTILINE|wx.TE_READONLY)
        box.Add(ff,1,wx.ALIGN_CENTRE|wx.ALL,10)
        box.Add(self.read_text,1,wx.ALIGN_CENTRE|wx.ALL,10)
        pl.SetSizer(box)
 #绑定事件(finish)--------------------------------------------------------------------------------------------------------
        self.Bind(wx.EVT_BUTTON,self.start_server,server_btn)
        self.Bind(wx.EVT_BUTTON,self.stop_server,stop_server_btn)
 #定义事件()--------------------------------------------------------------------------------------------------------------
    def send_message(self,socket,data):
        socket.send(data.encode('utf-8'))
    def recv_message(self,socket):
        return socket.recv(1024).decode('utf-8')
 #开始服务事件(finish)
    def start_server(self,event):
        if self.isON == False:
            self.isON = True
            self.read_text.AppendText("服务开启\n" + time.strftime('%Y-%m-%d %H:%M-%S', time.localtime()) + '\n')
            main_thread = Thread(target=self.start_work)
            main_thread.daemon = True
            main_thread.start()
 #主进程事件(done)
    def start_work(self):
        self.listen_socket = socket(AF_INET, SOCK_STREAM)
        self.listen_socket.bind(self.host_port)
        self.listen_socket.listen(5)
        while self.isON:
            try:
                session_socket,client_addr = self.listen_socket.accept()#接收一个套接字
                self.socket_connecting[session_socket] = ''
                session_Thread = session_thread(session_socket,self,self.work_sheet,self.work_book)
                session_Thread.daemon = True
                session_Thread.start()
            except OSError:
                break

    #停止服务事件（done）
    def stop_server(self,event):
        if self.isON == True:
            for session_socket in self.socket_connecting.keys():
                session_socket.send("shutdown".encode('utf-8'))
            self.read_text.AppendText("正在关闭服务...")
            while self.socket_connecting:
                pass
            self.isON = False
            self.listen_socket.close()
            self.read_text.AppendText("服务已关闭")


class session_thread(Thread):
    def __init__(self,session_socket,server,work_sheet,work_book):
        super().__init__()
        self.session_socket = session_socket
        self.server = server
        self.work_sheet = work_sheet
        self.work_book = work_book
        self.isON = True
        self.logged = False
    def run(self) -> None:
        self.send_message("欢迎使用本系统！")
        while self.isON:
            method = self.recv_message()
            if method == 'login':
                user_qx,user_name= self.user_login_varify()
                if user_name != None:
                    server.socket_connecting[self.session_socket] = user_name
                    #root用户
                    if user_qx == 'admin':
                        self.root_ctrl(user_name)
                    #高权限用户
                    elif user_qx == 'common':
                        self.common_ctrl(user_name)
                    #普通用户
                    elif user_qx == 'sudo':
                        self.sudo_ctrl(user_name)
            elif method == 'sign':
                self.sign_user()
            elif method !='':
                self.send_message("请先登录系统！")
    #不同用户操作（）------------------------------------------------------------------------------------------------------
    #root操作
    def root_ctrl(self,name):
        option = "查询信息\n更改权限\n强制下线\n关闭系统"#最外层循环，不用return来break
        option_list = option.split('\n')
        self.send_message(option)
        while self.logged:
            order = self.recv_message()
            if not self.is_exit(order):
                if order == option_list[0]:
                    server.read_text.AppendText(
                                    f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.ro0(name)
                if order == option_list[1]:
                    server.read_text.AppendText(
                                    f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.ro1(name)
                if order == option_list[2]:
                    server.read_text.AppendText(
                                    f"{time.strftime('%Y-%m-%d %H-%M-%S',time.localtime())}\n{name} 执行了 {order} 指令")
                    self.ro2(name)
                if order == option_list[3]:
                    server.read_text.AppendText(
                                    f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.ro3(name)

            else:
                break
        self.change_condition(name)
    #查询信息模块（done）
    def ro0(self,name):
        option = "用户信息\n在线用户\n\n按 return 返回上一级"
        self.send_message(option)
        option_list = option.split("\n")
        user_dict = {}
        max_row = self.work_sheet.max_row + 1
        for i in range(2,max_row):
            user_dict[self.work_sheet.cell(i, 1).value] = {}
            for j in range(2,11):
                user_dict[self.work_sheet.cell(i,1).value][self.work_sheet.cell(1,j).value] = self.work_sheet.cell(i,j).value
        self.send_message(option)
        while True:
            order = self.recv_message()
            if not self.is_exit(order):
                if order == option_list[0]:
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.ro0_0(user_dict,name)
                    if not self.logged:
                        break
                if order == option_list[1]:
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.ro0_1(user_dict)
                    if not self.logged:
                        break
                if order == 'return':
                    break
            else:
                break
    #查询用户信息（done）
    def ro0_0(self,user_dict,name):
        data = ''
        for user in user_dict.keys():
            data += f"{user}\n"
        while True:
            self.send_message(data + "\n按 return 返回上一级")
            order = self.recv_message()
            if not self.is_exit(order):
                if order in user_dict.keys():
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    user_intro = ''
                    for intro in user_dict[order].keys():
                        user_intro += f"{intro}   :   {user_dict[order][intro]}\n"
                    self.send_message(user_intro+"\n按 return 返回上一级")
                    order = self.recv_message()
                    if not self.is_exit(order):
                        pass
                    else:
                        break
                elif order == 'return':
                    break
            else:
                break
    #查询在线用户(done)
    def ro0_1(self,user_dict):
        data = ''
        for user in user_dict.keys():
            if user_dict[user]['用户状态'] == 'on':
                data += f"{user}\n"
        self.send_message(data+"\n按 return 返回上一级")
        order = self.recv_message()
        if not self.is_exit(order):
            pass
        else:
            pass
    #更改权限模块（done）
    def ro1(self,name):
        max_row = self.work_sheet.max_row + 1
        data = ''
        user_row_dict = {}
        for i in range(2, max_row):
            data += f"{self.work_sheet.cell(i, 1).value}   :   {self.work_sheet.cell(i, 3).value}\n"
            user_row_dict[self.work_sheet.cell(i, 1).value] = i
        data += "\n\n输入 （用户名） 为其更改权限\n（sudo->common,common->sudo）\n\n按 return 返回上一级"
        while True:
            self.send_message(data)
            order = self.recv_message()
            if not self.is_exit(order):
                if order in user_row_dict.keys():
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    if self.work_sheet.cell(user_row_dict[order],3).value == 'common':
                        self.work_sheet.cell(user_row_dict[order], 3).value = 'sudo'
                        self.work_book.save(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")
                        self.send_message("权限已更改！\n\n按 return 返回上一级")
                        order = self.recv_message()
                        if not self.is_exit(order):
                            pass
                        else:
                            break
                    elif self.work_sheet.cell(user_row_dict[order],3).value == 'sudo':
                        self.work_sheet.cell(user_row_dict[order], 3).value = 'common'
                        self.work_book.save(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")
                        self.send_message("权限已更改！\n\n按 return 返回上一级")
                        order = self.recv_message()
                        if not self.is_exit(order):
                            pass
                        else:
                            break
                elif order == 'return':
                    break
            else:
                break
    #下线用户模块（done）
    def ro2(self,name):
        while True:
            on_user_list = []
            data = ''
            max_row = self.work_sheet.max_row + 1
            for i in range(2,max_row):
               if self.work_sheet.cell(i,8).value == 'on':
                    on_user_list.append(self.work_sheet.cell(i,1).value)
                    data += f"{self.work_sheet.cell(i, 1).value}  is  on\n"
            data += "\n\n按 return 返回上一级"
            self.send_message(data)
            order = self.recv_message()
            if not self.is_exit(order):
                if order in on_user_list:
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    for session in server.socket_connecting.keys():
                        if server.socket_connecting[session] == order:
                            session.send("exit-s".encode('utf-8'))
                            self.send_message("已关闭该用户！\n\n按 return 返回上一级")
                            order = self.recv_message()
                            if not self.is_exit(order):
                                pass
                            else:
                                break
                elif order == 'return':
                    break
            else:
                break
    #关闭系统模块
    def ro3(self,name):
        for session in server.socket_connecting.keys():
            if server.socket_connecting[session] != name:
                session.send("shutdown".encode("utf-8"))
        if len(server.socket_connecting) == 1:
            server.read_text.AppendText("其余客户机已关闭！！！\n\n\n按 return 返回上一级")
        order = self.recv_message()
        if not self.is_exit(order):
            pass
        else:
            pass


    #sudo用户操作
    def sudo_ctrl(self,name):
        rows = self.work_sheet.max_row + 1
        for i in range(2, rows):
            if self.work_sheet.cell(i, 1).value == name:
                user_row = i
                break
        option = "查询信息\n打印本月财务报表"
        option_list = option.split("\n")
        while self.logged:
            self.send_message(option)
            order = self.recv_message()
            if not self.is_exit(order):
                if order == option_list[0]:
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.so0(user_row,name,rows)
                elif order == option_list[1]:
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.so1(user_row)
            else:
                break
        self.change_condition(name)
    def so0(self,row,name,rows):
        option = "查询本用户信息\n查询common用户密码"
        option_list = option.split("\n")
        while True:
            self.send_message(option+"\n\n按 return 返回上一级")
            order = self.recv_message()
            if not self.is_exit(order):
                if order == option_list[0]:
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.so0_1(row, name)
                    if not self.logged:
                        break
                elif order == option_list[1]:
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.so0_2(rows, name)
                    if not self.logged:
                        break
                elif order == 'return':
                    break
            else:
                break
    def so0_1(self,row,name):
        option_dict = {}
        data1 = ''
        for i in range(1, 11):
            option_dict[self.work_sheet.cell(1, i).value] = self.work_sheet.cell(row, i).value
            data1 += f"{self.work_sheet.cell(1, i).value}\n"
        data1 += "\n按 return 返回上一级"
        while True:
            self.send_message(data1)
            order = self.recv_message()
            if not self.is_exit(order):
                if order == 'return':
                    break
                if order in option_dict.keys():
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    data = str(option_dict[order])
                    self.send_message(data+"\n\n按 return 返回上一级")
                    order = self.recv_message()
                    if not self.is_exit(order):
                        pass
                    else:
                        break
            else:
                break
    def so0_2(self,rows,name):
        com_passwd_dict = {}
        send_data = ''
        for i in range(3,rows):
            if self.work_sheet.cell(i,3).value == 'common':
                com_passwd_dict[self.work_sheet.cell(i,1).value] = self.work_sheet.cell(i,2).value
                send_data += f"{self.work_sheet.cell(i,1).value}\n"
        while True:
            self.send_message(send_data+"\n返回上一级请按-------return")
            order = self.recv_message()
            if not self.is_exit(order):
                if order == 'return':
                    break
                if order in com_passwd_dict.keys():
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.send_message(f"{order}  的密码：  {com_passwd_dict[order]}\n\n按 return 返回上一级")
                    order = self.recv_message()
                    if not self.is_exit(order):
                        pass
                    else:
                        break
            else:
                break
    def so1(self,row):
        self.send_message(f"{self.work_sheet.cell(row, 9).value}\n\n按 return 返回上一级")
        order = self.recv_message()
        if not self.is_exit(order):
            pass
        else:
            pass



    #普通用户操作（done）
    def common_ctrl(self,name):
        #找到行数
        rows  = self.work_sheet.max_row + 1
        for i in range(3,rows):
            if self.work_sheet.cell(i,1).value == name:
                user_row = i
                break
        while self.logged:
            option = "查询信息\n打印本月财务报表"
            option_list =option.split("\n")
            self.send_message(option)
            order = self.recv_message()
            if not self.is_exit(order):
                if order == option_list[0]:
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.co0(user_row,name)
                elif order == option_list[1]:
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    self.co1(user_row)
                elif not order:
                    break
        self.change_condition(name)
    def co0(self,row,name):
        option_dict = {}
        data = ''
        for i in range(1,8):
            option_dict[self.work_sheet.cell(1,i).value] = self.work_sheet.cell(row,i).value
            data += f"{self.work_sheet.cell(1,i).value}\n"
        while True:
            self.send_message(data+"\n 返回上一级请按------return")
            order = self.recv_message()
            if not self.is_exit(order):
                if order == 'return':
                    break
                elif order in option_dict.keys():
                    server.read_text.AppendText(
                        f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{name} 执行了 {order} 指令")
                    respon = str(option_dict[order])
                    self.send_message(respon+"\n\n按 return 返回上一级")
                    order = self.recv_message()
                    if not self.is_exit(order):
                        pass
                    else:
                        break
            else:
                break
    def co1(self,row):
        self.send_message(f"{self.work_sheet.cell(row, 9).value}\n\n按 return 返回上一级")
        order = self.recv_message()
        if not self.is_exit(order):
            pass
        else:
            pass
    #定义简单函数----------------------------------------------------------------------------------------------------------
    # 检测用户名，如果正确就返回用户名与权限
    def user_login_varify(self):
        user_name = self.recv_message()
        passwd = int(self.recv_message())
        user_passwd_dict = {}
        user_qx_dict = {}
        user_rows = self.work_sheet.max_row
        for i in range(2, user_rows + 1):
            user_passwd_dict[self.work_sheet.cell(i, 1).value] = self.work_sheet.cell(i, 2).value  # 数字类型都是int
            user_qx_dict[self.work_sheet.cell(i, 1).value] = self.work_sheet.cell(i, 3).value
        if user_name in user_passwd_dict.keys():
            if passwd == user_passwd_dict[user_name]:
                self.send_message("logged")
                server.read_text.AppendText(f"{time.strftime('%Y-%m-%d %H-%M-%S',time.localtime() )}\n{user_name}  登录进系统")
                self.change_condition(user_name)
                self.logged = True
                return user_qx_dict[user_name], user_name
            else:
                self.send_message("密码错误！")
                return None, None
        else:
            self.send_message("用户不存在！")
            return None, None
    def change_condition(self,user):
        rows = self.work_sheet.max_row + 1
        for i in range(2,rows):
            if self.work_sheet.cell(i,1).value == user:
                if self.work_sheet.cell(i,8).value == 'down':
                    self.work_sheet.cell(i,8).value = 'on'
                else:
                    self.work_sheet.cell(i, 8).value = 'down'
        self.work_book.save(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")

    #定义接受信息的函数
    def recv_message(self):
        if self.isON:
            data = self.session_socket.recv(1024).decode('utf-8')
            if not data:
                self.isON = False
                del server.socket_connecting[self.session_socket]
                self.session_socket.close()
            return data



    #定义发信息的函数
    def send_message(self,data):
        if self.isON:
            self.session_socket.send(data.encode('utf-8'))
    #定义注册用户(done)
    def sign_user(self):
        rows = self.work_sheet.max_row + 1
        user_dict = []
        for i in range(3,rows):
            user_dict.append(self.work_sheet.cell(i,1).value)
        self.send_message("请输入注册的用户名")
        user_name = self.recv_message()
        while user_name in user_dict:
            self.send_message("该用户已存在！")
            user_name = self.recv_message()
        self.send_message("请输入密码")
        passwd = self.recv_message()
        while passwd == '':
            self.send_message("密码不能为空！")
            passwd = self.recv_message()
        self.send_message("请输入性别")
        sex = self.recv_message()
        while sex != '男' and sex != '女' and sex != 'female' and sex != 'male':
            self.send_message("只能填写（男/女/female/male）")
            sex = self.recv_message()
        self.send_message("请输入居住地")
        address = self.recv_message()
        while address == '':
            self.send_message("居住地不能为空！")
            address = self.recv_message()
        self.send_message("请输入生日（1111-11-11）")
        date = self.recv_message()
        while date == '':
            self.send_message("生日不能为空！")
            date = self.recv_message()
        qx = 'common'
        id = self.work_sheet.max_row
        bb = 'dads'
        user_cond = 'down'
        sign_time = time.strftime("%Y-%m-%d",time.localtime())
        new_user_list =[user_name,passwd,qx,sex,date,address,sign_time,user_cond,bb,id]
        for i in range(1,11):
            self.work_sheet.cell(rows,i).value = new_user_list[i-1]
        self.work_book.save(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")
        self.send_message("用户已创建！")
        server.read_text.AppendText(f"{time.strftime("%Y-%m-%d %H:%M-%S", time.localtime())}\n{user_name} 已被创建 " )
        self.send_message("signned")
    def is_exit(self,order):
        if order == "exit-c":
            self.logged = False
            self.send_message(order)
            return True
        else:
            return False

#主程序运行---------------------------------------------------------------------------------------------------------------
if __name__ == '__main__':
    app = wx.App()
    server = server_answer(work_sheet,work_book)
    server.Show()
    app.MainLoop()