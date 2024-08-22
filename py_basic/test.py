import openpyxl
from socket import AF_INET,socket,SOCK_STREAM
import time

data = openpyxl.load_workbook(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")
work_sheet = data.active # 目前工作目录
work_sheet.cell(2, 8).value = 'on'
data.save(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")



user_dict = {}
max_row = work_sheet.max_row + 1
for i in range(2, max_row):
    user_dict[work_sheet.cell(i, 1).value] = {}
    for j in range(2, 11):
        user_dict[work_sheet.cell(i, 1).value][work_sheet.cell(1, j).value] = work_sheet.cell(i, j).value




# for user in user_dict.keys():
#     print(user_dict[user]['用户状态'])
    # for intro in user_dict[user].keys():
    #     # print(user_dict[user][intro])
    # break
# print(user_dict.keys())
# work_sheet.cell(1,1).value = time.strftime("%Y-%m-%d",time.localtime())
# data.save(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表 (2).xlsx")
# print(work_sheet.cell(1,1).value)
# print(work_sheet.cell(2,1).value)
# print(type(work_sheet.cell(1,1).value))
# print(type(work_sheet.cell(2,1).value))
# data.save(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表 (2).xlsx")
# print(type(work_sheet.cell(2,5).value))
#
# print(work_sheet.max_row)
# user_list = []
# for i in range(2,int(work_sheet.max_row)+1):
#     print(work_sheet.cell(2,i).value)
# print(type(work_sheet.max_row + 1))
# work_sheet['A11'].value = 'qiqisxd'
# data.save(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")
#
# work_sheet['A18'].value = 'qiqcdisxd'
# data.save(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")
# work_sheet['A19'].value = 'qiqisxd'
# data.save(r"C:\Users\smart dogs\Desktop\新建 XLSX 工作表.xlsx")
# import time
#
# data = 2
# while (data != 1 and data != 2):
#     print("aaa")
#     time.sleep(1)
# list = "\nw\nx\nw\nw\nw\n"
# print(list.split('\n'))


# def ddd():
#     return None
#
# data = ddd(
#
#
# server_listen_socket = socket(AF_INET,SOCK_STREAM)
# server_listen_socket.bind(('',8888))
# server_listen_socket.listen(1)

# dcit = {'daf':'dwfq','affa':'fef'}
# for i in dcit:
#     print(i)
# while True:
#     session_socket, addr = server_listen_socket.accept()
#     isON = True
#     while isON:
#         data = session_socket.recv(1024).decode('utf-8')
#         if data != 'exit' and data:
#             print(data)
#         elif data == 'exit':
#             session_socket.send("已退出登录".encode('utf-8'))
#         else:
#             session_socket.close()
#             print("结束对话")
#             break



# option = "查询信息\n打印本月财务报表"
# option_list =option.split("\n")
# print(type(option_list.index('查询信息')))

